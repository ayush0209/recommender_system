import numpy as np
import multiprocessing
import math
from scipy import sparse 
from openpyxl import load_workbook

npr=4
#function for predicting rating of a particular item for a user using user-user collaborative filtering
def rating_calculator(unq,ro,co,rat,spars_matrix,spars_matrix_transpose,max_row_index,original_matrix,lock,us,ac,pre,usrin,actin,precin):

	xx=int(unq*(len(ro)/npr))
	yy=int((unq+1)*(len(ro)/npr))
	
	#converting to dense matrix
	spars_matrix=spars_matrix.toarray()
	spars_matrix_transpose=spars_matrix_transpose.toarray()
	sim_score_matrix_dense=np.dot(spars_matrix,spars_matrix_transpose)
	original_matrix_dense=original_matrix.toarray()
	diff=[]

	for j in range(xx,yy):
		row_index=int(ro[j])
		col_index=int(co[j])

		weighted_average=-1
		prod=0
		sum_of_weights=0
		sim_score_list={}

		dx=0

		for i in range(max_row_index):
			ith_row_sim_score=0
			if(i!=row_index):
				if(original_matrix_dense[i][col_index]>=1):
					if(sim_score_matrix_dense[row_index][i]>=0):
						ith_row_sim_score=sim_score_matrix_dense[row_index][i]
						dx+=1
				sim_score_list[i]=ith_row_sim_score	
			
		sorted_by_value=sorted(sim_score_list.items(), key=lambda kv: kv[1], reverse=True)

		for y in range(min(10,dx)):

			rowy=sorted_by_value[y][0]
			weight=sorted_by_value[y][1]
			nthg=original_matrix_dense[rowy][col_index]

			sum_of_weights+=weight
			prod+=(nthg*weight)

		if(sum_of_weights!=0):
			weighted_average=prod/sum_of_weights
			diff.append(rat[j]-weighted_average)

		#writing to shared memory
		lock.acquire()

		us[usrin.value]=row_index
		usrin.value+=1
		ac[actin.value]=rat[j]
		actin.value+=1
		pre[precin.value]=weighted_average
		precin.value+=1

		lock.release()
		#calculating RMSE
	RMSE=math.sqrt(sum(i**2 for i in diff)/len(diff))
	#writing RMSE calculated by process back to excel file

	lock.acquire()
	wb=load_workbook("answer.xlsx")
	ws=wb.active
	
	cell1="A"+str(unq+1)
	cell2="B"+str(unq+1)	

	ws[cell1]=RMSE
	ws[cell2]=len(diff)
	wb.save("answer.xlsx")

	lock.release()